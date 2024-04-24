from typing import List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from addict import Dict


def is_empty_entry(ws: Worksheet, row, column):
    val = ws.cell(row=row, column=column).value
    if val is None or str(val).strip() == '':
        return True
    return False


def get_num_rows_cols(ws: Worksheet,
                      header_row_idx=1, non_empty_col=1):
    """
    This function determines the actual number of rows and columns
        in an Excel worksheet based on the column headers and
        entries that are supposed to be not empty
    """
    max_row, max_col = int(ws.max_row), int(ws.max_column)
    # first determine the number of rows through binary search
    min_row = 1
    while min_row < max_row:
        curr_row = (min_row + max_row) // 2
        if is_empty_entry(ws, curr_row, non_empty_col):
            max_row = curr_row - 1
        else:
            # plus one ensures min_row and max_row could eventually equal to each other
            min_row = curr_row + 1
    num_row = min_row if not is_empty_entry(ws, min_row, non_empty_col) else min_row - 1

    # then determine the number of columns through binary search
    min_col = 1
    while min_col < max_col:
        curr_col = (min_col + max_col) // 2
        if is_empty_entry(ws, row=header_row_idx, column=curr_col):
            max_col = curr_col - 1
        else:
            # plus one ensures min_col and max_col could eventually equal to each other
            min_col = curr_col + 1
    num_col = min_col if not is_empty_entry(ws, header_row_idx, min_col) else min_col - 1

    return Dict({'num_row': num_row, 'num_col': num_col})


def get_column2idx(ws: Worksheet, header_row_idx=1, num_column=None):
    if num_column is None:
        num_column = get_num_rows_cols(ws, header_row_idx=header_row_idx)['num_col']
    column2idx = Dict()
    # noinspection PyTypeChecker
    for col_idx in range(1, num_column+1):
        col_name = str(ws.cell(row=header_row_idx, column=col_idx).value).strip()
        column2idx[col_name] = col_idx
    return column2idx


def read_rows(ws: Worksheet,
              rows_to_read: List, column2idx=None,
              header_row_idx=1, num_column=None):
    if num_column is None:
        num_column = get_num_rows_cols(ws, header_row_idx=header_row_idx)['num_col']
    if column2idx is None:
        column2idx = get_column2idx(ws, header_row_idx=header_row_idx, num_column=num_column)

    data = list()
    for row_idx in rows_to_read:
        row_data = Dict()
        for col_name, col_idx in column2idx.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            val = '' if val is None else val
            row_data[col_name] = val
        data.append(row_data)
    return data


class ExcelReader:
    def __init__(self, excel_file_path, sheet_name,
                 header_row_idx=1, non_empty_col=1):
        """
        The non_empty_col corresponds to a field where each data entry should have a non-empty value (this is
            used to determine the number of rows in an Excel sheet)
        """
        self.excel_file_path = excel_file_path
        self.sheet_name = sheet_name

        self.wb = openpyxl.load_workbook(excel_file_path)
        self.ws = self.wb[sheet_name]
        self.is_wb_open = True

        # init relevant variables
        self.header_row_idx, self.non_empty_col = header_row_idx, non_empty_col
        self.n_col, self.column2idx = None, None
        self._init_col_info()

    def _reopen_workbook(self):
        self.wb = openpyxl.load_workbook(self.excel_file_path)
        self.ws = self.wb[self.sheet_name]
        self._init_col_info()
        self.is_wb_open = True

    def _check_workbook_open(self):
        if not self.is_wb_open:
            self._reopen_workbook()

    def _init_col_info(self):
        n_row_col = get_num_rows_cols(self.ws, header_row_idx=self.header_row_idx, non_empty_col=self.non_empty_col)
        self.n_col = n_row_col['num_col']
        self.column2idx = get_column2idx(self.ws, header_row_idx=self.header_row_idx, num_column=self.n_col)

    def get_rows_data(self, rows, append_row_idx=False):
        self._check_workbook_open()
        rows_data = read_rows(self.ws, rows, column2idx=self.column2idx,
                              header_row_idx=self.header_row_idx, num_column=self.n_col)
        if append_row_idx:
            for i, row_idx in enumerate(rows):
                rows_data[i]['row_idx'] = row_idx
        return rows_data

    def n_rows(self):
        self._check_workbook_open()
        return get_num_rows_cols(self.ws, header_row_idx=self.header_row_idx,
                                 non_empty_col=self.non_empty_col)['num_row']

    def update_row(self, row_idx, value, col_name=None, col_idx=None):
        """
        This function updates the data of the row with index row_idx. The user needs
            to specify either col_name or col_idx. When adding a new column, col_name
            should be specified
        """
        self._check_workbook_open()
        if col_idx is None:
            assert col_name is not None
            if col_name in self.column2idx:
                col_idx = self.column2idx[col_name]
            else:
                # this is a new column
                self.n_col += 1
                self.column2idx[col_name] = self.n_col
                col_idx = self.n_col
                # update the header
                self.ws.cell(row=self.header_row_idx,
                             column=col_idx).value = col_name
        self.ws.cell(row=row_idx, column=col_idx).value = value

    def save(self, file_path=None):
        self._check_workbook_open()
        file_path = self.excel_file_path if file_path is None else file_path
        self.wb.save(filename=file_path)
        self.wb.close()
        self.is_wb_open = False







